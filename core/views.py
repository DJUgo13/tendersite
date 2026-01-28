from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.utils import timezone
from django.core.mail import send_mail, EmailMultiAlternatives
from django.conf import settings
from django.template.loader import render_to_string
from django.http import HttpResponse, JsonResponse, StreamingHttpResponse
from datetime import timedelta
import json
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from .models import Direction, Bid, Tender, Company, UserProfile, AuditLog, Winner
from .utils import (
    log_action,
    get_user_role,
    is_admin,
    check_and_handle_rebidding,
    get_anonymous_best_price,
    check_auto_close_directions,
    update_final_timer,
    send_tender_started_emails,
    send_winner_email,
)
from .decorators import admin_required, manager_required, leadership_required


def login_view(request):
    """Страница входа в систему с проверкой наличия привязанной Company"""
    if request.user.is_authenticated:
        role = get_user_role(request.user)
        # Автоматическая переадресация при наличии сессии
        if role == 'admin' or role == 'leadership':
            return redirect('leadership_dashboard')
        else:
            return redirect('manager_dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            
            # Логируем вход
            log_action(user, 'user_login', 'user', user.id, {'username': username}, request)
            
            role = get_user_role(user)
            
            # Проверяем роль и перенаправляем
            if role == 'admin':
                # Администраторы также попадают в панель руководства по умолчанию
                # Доступ к админке есть в меню
                return redirect('leadership_dashboard')
            elif role == 'leadership':
                return redirect('leadership_dashboard')
            else:
                # Для менеджеров проверяем наличие компании
                try:
                    company = user.company
                    return redirect('manager_dashboard')
                except Company.DoesNotExist:
                    messages.error(request, 'У вашего аккаунта нет привязанной компании. Обратитесь к администратору.')
                    logout(request)
                    return render(request, 'core/login.html')
        else:
            messages.error(request, 'Неверный логин или пароль.')
    
    return render(request, 'core/login.html')


def switch_account(request):
    """Выход и переход на страницу входа для смены аккаунта."""
    if request.user.is_authenticated:
        log_action(request.user, 'user_logout', 'user', request.user.id, {'reason': 'switch_account'}, request)
        logout(request)
    return redirect('login')


def logout_view(request):
    """Выход из системы"""
    if request.user.is_authenticated:
        log_action(request.user, 'user_logout', 'user', request.user.id, {}, request)
    logout(request)
    messages.success(request, 'Вы успешно вышли из системы.')
    return redirect('login')


@manager_required
def manager_dashboard(request):
    """Главная страница менеджера - список открытых тендеров для текущей компании"""
    # Проверяем наличие компании у пользователя
    try:
        user_company = request.user.company
    except Company.DoesNotExist:
        messages.error(request, 'У вашего аккаунта нет привязанной компании.')
        logout(request)
        return redirect('login')
    
    # Получаем все открытые тендеры
    tenders = Tender.objects.filter(status='open').order_by('-created_at')
    return render(request, 'core/manager_dashboard.html', {'tenders': tenders})


def help_page(request):
    """Страница помощи"""
    return render(request, 'core/help.html')


@manager_required
def tender_detail(request, tender_id):
    """
    Детальная страница тендера с направлениями.
    Показывает:
    - текущую лучшую цену
    - ставку текущей компании
    - после завершения тендера — информацию о победе текущей компании
    - таймер до окончания торгов
    """
    tender = get_object_or_404(Tender, id=tender_id)

    # Автоматическое завершение тендера по времени:
    # если время окончания прошло, а статус ещё "open" – закрываем, определяем победителей и отправляем письма
    if tender.status == 'open' and tender.end_time and tender.end_time <= timezone.now():
        # Закрываем тендер и отправляем письма победителям
        results = close_tender_and_notify_winners(tender, set_end_time=False)
        
        # Логируем результаты отправки (можно добавить в messages, но это может быть слишком много)
        # Для автоматического закрытия просто фиксируем результаты

    # Подстраховка: закрываем направления по финальному таймеру (если команда/cron не настроены)
    check_auto_close_directions()

    directions = tender.directions.all()
    
    # Получаем компанию пользователя
    try:
        user_company = request.user.company
    except Company.DoesNotExist:
        messages.error(request, 'У вашего аккаунта нет привязанной компании.')
        return redirect('manager_dashboard')
    
    user_won_any = False
    user_won_directions = []

    # Для каждого направления добавляем информацию о ставке пользователя и текущей лучшей цене
    for direction in directions:
        # Текущая лучшая цена (уже есть в модели, но убедимся что она актуальна)
        if not direction.current_best_price:
            # Если лучшая цена не установлена, используем стартовую
            direction.current_best_price = direction.start_price
        
        # Ставка текущей компании
        bid = (
            Bid.objects.filter(tender=tender, direction=direction, company=user_company, is_active=True)
            .order_by('-created_at')
            .first()
        )
        direction.my_bid_price = bid.price if bid else None

        # Флаг: является ли текущая компания победителем по направлению
        direction.is_winner_for_me = (
            tender.status == 'closed'
            and direction.winner_id == user_company.id
        )

        if direction.is_winner_for_me:
            user_won_any = True
            user_won_directions.append(direction)
    
    return render(request, 'core/tender_detail.html', {
        'tender': tender,
        'directions': directions,
        'user_won_any': user_won_any,
        'user_won_directions': user_won_directions,
    })


@manager_required
def get_best_price(request, direction_id):
    """API endpoint для получения лучшей цены по направлению (для HTMX)"""
    direction = get_object_or_404(Direction, id=direction_id)
    return render(request, 'core/partials/best_price.html', {
        'best_price': direction.current_best_price
    })


# Обработка отправки ставки (форма из модального окна)
@manager_required
def submit_bid(request):
    """Обработка POST-запроса на создание/обновление ставки с валидацией"""
    if request.method != 'POST':
        return redirect('manager_dashboard')
    
    direction_id = request.POST.get('direction_id')
    price = request.POST.get('price')

    if not direction_id or not price:
        messages.error(request, 'Не все поля заполнены.')
        referer = request.META.get('HTTP_REFERER')
        if referer:
            return redirect(referer)
        return redirect('manager_dashboard')

    direction = get_object_or_404(Direction, id=direction_id)
    
    # Получаем компанию пользователя
    try:
        user_company = request.user.company
    except Company.DoesNotExist:
        messages.error(request, 'У вашего аккаунта нет привязанной компании.')
        return redirect('manager_dashboard')

    # Проверка: тендер должен быть открыт
    if direction.tender.status != 'open':
        messages.error(request, 'Тендер закрыт, ставки не принимаются.')
        return redirect('tender_detail', tender_id=direction.tender.id)

    # Проверка: цена должна быть числом и выше 0
    try:
        price = float(price)
        if price <= 0:
            raise ValueError
    except (ValueError, TypeError):
        messages.error(request, 'Введите корректную цену (положительное число).')
        return redirect('tender_detail', tender_id=direction.tender.id)

    now = timezone.now()

    # Если направление уже "закрыто" по таймеру (winner выставлен) — запрещаем ставки
    if direction.winner_id and direction.final_timer_end and direction.final_timer_end <= now and direction.tender.status == 'open':
        messages.error(request, 'Направление закрыто по финальному таймеру. Ставки больше не принимаются.')
        return redirect('tender_detail', tender_id=direction.tender.id)

    # Лучший активный бид (для корректной проверки равенства/переторжки)
    best_bid = direction.bids.filter(is_active=True).order_by('price', 'created_at').first()
    current_best = best_bid.price if best_bid else (direction.current_best_price or direction.start_price)

    # Если идёт переторжка — принимать ставки только от компаний, которые сейчас в минимуме
    if direction.is_in_rebidding and direction.rebidding_end_time and direction.rebidding_end_time > now:
        min_price = direction.bids.filter(is_active=True).order_by('price', 'created_at').first().price
        tied_company_ids = list(
            direction.bids.filter(is_active=True, price=min_price).values_list('company_id', flat=True)
        )
        if user_company.id not in tied_company_ids:
            messages.error(request, 'Сейчас идёт переторжка между лидерами. Ваша компания не участвует в переторжке.')
            return redirect('tender_detail', tender_id=direction.tender.id)
    elif direction.is_in_rebidding and direction.rebidding_end_time and direction.rebidding_end_time <= now:
        # Переторжка закончилась — сбрасываем флаг
        direction.is_in_rebidding = False
        direction.rebidding_end_time = None
        direction.save(update_fields=['is_in_rebidding', 'rebidding_end_time'])

    # Правило: ставка должна быть НЕ ВЫШЕ текущей лучшей.
    # Равная цена допускается, чтобы запустить переторжку при равенстве.
    if price > float(current_best):
        messages.error(
            request,
            f'Ваша ставка ({price:.2f} руб.) должна быть не выше текущей лучшей ({float(current_best):.2f} руб.).'
        )
        return redirect('tender_detail', tender_id=direction.tender.id)

    # Деактивируем старую ставку, если она есть
    old_bid = Bid.objects.filter(
        tender=direction.tender,
        direction=direction,
        company=user_company,
        is_active=True
    ).first()
    
    if old_bid:
        old_bid.is_active = False
        old_bid.save()
    
    # Создаём новую ставку
    bid = Bid.objects.create(
        tender=direction.tender,
        direction=direction,
        company=user_company,
        price=price,
        created_by=request.user,
        is_active=True
    )

    # Обновляем лучшую цену в направлении (новая ставка всегда лучше, т.к. прошла валидацию)
    direction.current_best_price = price
    direction.save()
    
    # Обновляем финальный таймер
    update_final_timer(direction)
    
    # Проверяем переторжку
    check_and_handle_rebidding(direction)

    # Логируем действие
    action = 'bid_created' if not old_bid else 'bid_updated'
    log_action(request.user, action, 'bid', bid.id, {
        'tender_id': direction.tender.id,
        'direction_id': direction.id,
        'price': float(price),
        'company': user_company.name
    }, request)

    if not old_bid:
        messages.success(request, f'Ставка {price:.2f} руб. на направление {direction.city_name} успешно размещена!')
    else:
        messages.success(request, f'Ваша ставка на направление {direction.city_name} обновлена до {price:.2f} руб.!')

    return redirect('tender_detail', tender_id=direction.tender.id)


@manager_required
def sse_tender_prices(request, tender_id):
    """
    SSE поток с лучшими ценами по направлениям тендера.
    Клиент получает JSON со списком направлений раз в ~3 секунды.
    """
    tender = get_object_or_404(Tender, id=tender_id)

    def event_stream():
        # Первый "hello" пакет
        yield "retry: 3000\n"
        yield f"data: {json.dumps({'type': 'hello', 'tender_id': tender.id})}\n\n"

        while True:
            # Подстраховка таймеров
            check_auto_close_directions()

            payload = {
                'type': 'prices',
                'tender_id': tender.id,
                'ts': timezone.now().isoformat(),
                'directions': []
            }
            for d in tender.directions.all():
                payload['directions'].append({
                    'id': d.id,
                    'best_price': str(d.current_best_price or d.start_price),
                    'is_in_rebidding': bool(d.is_in_rebidding),
                    'rebidding_end_time': d.rebidding_end_time.isoformat() if d.rebidding_end_time else None,
                    'final_timer_end': d.final_timer_end.isoformat() if d.final_timer_end else None,
                    'winner_id': d.winner_id,
                })

            yield f"data: {json.dumps(payload)}\n\n"
            # Каждые 3 секунды (простая реализация SSE)
            import time
            time.sleep(3)

    response = StreamingHttpResponse(event_stream(), content_type='text/event-stream')
    response['Cache-Control'] = 'no-cache'
    return response





def close_tender_and_notify_winners(tender, set_end_time=True, request=None):
    """
    Закрывает тендер, определяет победителей и отправляет им письма.
    Возвращает список результатов отправки писем.
    """
    if tender.status != 'open':
        return []
    
    # Меняем статус
    tender.status = 'closed'
    if set_end_time and not tender.end_time:
        tender.end_time = timezone.now()
    tender.save()
    
    # Логируем закрытие тендера
    if request:
        log_action(request.user if request.user.is_authenticated else None, 
                  'tender_closed', 'tender', tender.id, {}, request)
    
    # Определяем победителей по каждому направлению и отправляем письма
    results = []
    for direction in tender.directions.all():
        # Находим минимальную ставку среди активных
        winning_bid = direction.bids.filter(is_active=True).order_by('price', 'created_at').first()
        
        if winning_bid:
            # Записываем победителя
            direction.winner = winning_bid.company
            direction.final_price = winning_bid.price
            direction.current_best_price = winning_bid.price
            direction.save()
            
            # Создаём запись в Winner для истории
            Winner.objects.update_or_create(
                tender=tender,
                direction=direction,
                defaults={
                    'company': winning_bid.company,
                    'price': winning_bid.price
                }
            )
            
            # Отправляем email победителю
            success, message = send_winner_email(tender, direction, winning_bid)
            results.append({
                'direction': direction,
                'winner': winning_bid.company,
                'success': success,
                'message': message
            })
    
    return results


# Закрытие тендера (только для админов)
@staff_member_required
def close_tender(request, tender_id):
    tender = get_object_or_404(Tender, id=tender_id)
    
    if tender.status != 'open':
        messages.error(request, 'Тендер уже закрыт или не был открыт.')
        return redirect('admin:core_tender_changelist')
    
    # Закрываем тендер и отправляем письма победителям
    results = close_tender_and_notify_winners(tender, set_end_time=True, request=request)
    
    # Выводим сообщения о результатах отправки писем
    successful_emails = sum(1 for r in results if r['success'])
    failed_emails = len(results) - successful_emails
    
    if settings.EMAIL_BACKEND != 'django.core.mail.backends.smtp.EmailBackend':
        messages.info(request, 'Примечание: Система работает в режиме отладки (SMTP не настроен). Письма выведены в консоль сервера.')
    
    for result in results:
        if result['success']:
            messages.success(request, result['message'])
        else:
            messages.warning(request, result['message'])
    
    messages.success(request, f'Тендер "{tender.name}" успешно закрыт. Определено победителей: {len(results)}. Писем отправлено: {successful_emails}, ошибок: {failed_emails}')
    return redirect('admin:core_tender_changelist')


@staff_member_required
def open_tender(request, tender_id):
    """Открытие тендера (ручной запуск) + уведомления о старте"""
    tender = get_object_or_404(Tender, id=tender_id)

    if tender.status == 'open':
        messages.info(request, 'Тендер уже открыт.')
        return redirect('admin:core_tender_changelist')

    if tender.status == 'closed':
        messages.error(request, 'Тендер уже закрыт. Открыть повторно нельзя.')
        return redirect('admin:core_tender_changelist')

    tender.status = 'open'
    if not tender.start_time:
        tender.start_time = timezone.now()
    tender.save()

    log_action(request.user, 'tender_opened', 'tender', tender.id, {'name': tender.name}, request)

    results = send_tender_started_emails(tender)
    ok = sum(1 for r in results if r['success'])
    fail = len(results) - ok
    
    is_smtp = settings.EMAIL_BACKEND == 'django.core.mail.backends.smtp.EmailBackend'
    msg_type = 'отправлено' if is_smtp else 'выведено в консоль (SMTP не настроен)'
    
    messages.success(request, f'Тендер "{tender.name}" успешно открыт.')
    messages.info(request, f'Рассылка ({msg_type}): успешно {ok}, ошибок {fail}.')
    
    if not settings.EMAIL_CONFIGURED and settings.EMAIL_MODE == 'production':
        messages.error(request, 'КРИТИЧНО: Почта не настроена в settings.py! Письма не ушли.')

    # Показываем детальный лог (до 10 записей)
    for r in results[:10]:
        icon = "✅" if r['success'] else "❌"
        messages.debug(request, f"{icon} {r['message']}")
        if not r['success']:
            messages.warning(request, f"Ошибка отправки на {r['email']}: {r['message']}")

    return redirect('admin:core_tender_changelist')






def _build_tender_excel_response(tender):
    """Внутренняя функция генерации Excel отчёта (используется в админке и для руководства)."""
    # Создаём новую рабочую книгу
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Тендер {tender.name}"

    # Стили для заголовков
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Заголовок тендера
    ws.merge_cells('A1:F1')
    ws['A1'] = f'Тендер: {tender.name}'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

    ws['A2'] = f'Статус: {tender.get_status_display()}'
    ws['A3'] = f'Дата создания: {tender.created_at.strftime("%d.%m.%Y %H:%M")}'
    if tender.end_time:
        ws['A4'] = f'Дата закрытия: {tender.end_time.strftime("%d.%m.%Y %H:%M")}'

    # Заголовки таблицы
    row = 6
    headers = ['Направление', 'Город', 'Объём (машин)', 'Стартовая цена', 'Финальная цена', 'Победитель']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Данные по направлениям
    row = 7
    for direction in tender.directions.all():
        ws.cell(row=row, column=1).value = direction.city_name
        ws.cell(row=row, column=2).value = direction.city_name
        ws.cell(row=row, column=3).value = direction.volume
        ws.cell(row=row, column=4).value = float(direction.start_price)

        if direction.final_price and direction.winner:
            ws.cell(row=row, column=5).value = float(direction.final_price)
            ws.cell(row=row, column=6).value = direction.winner.name
        else:
            ws.cell(row=row, column=5).value = 'Нет победителя'
            ws.cell(row=row, column=6).value = '-'

        row += 1

    # Настройка ширины столбцов
    column_widths = [20, 20, 15, 15, 15, 30]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Форматирование чисел
    for r in range(7, row):
        ws.cell(row=r, column=4).number_format = '#,##0.00'
        if ws.cell(row=r, column=5).value and isinstance(ws.cell(row=r, column=5).value, (int, float)):
            ws.cell(row=r, column=5).number_format = '#,##0.00'

    # Лист со всеми ставками
    ws2 = wb.create_sheet(title="Все ставки")

    headers2 = ['Направление', 'Компания', 'ИНН', 'Цена ставки', 'Дата ставки', 'Менеджер', 'Активна']
    for col, header in enumerate(headers2, start=1):
        cell = ws2.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    rr = 2
    for direction in tender.directions.all():
        for bid in direction.bids.all().order_by('price', 'created_at'):
            ws2.cell(row=rr, column=1).value = direction.city_name
            ws2.cell(row=rr, column=2).value = bid.company.name
            ws2.cell(row=rr, column=3).value = bid.company.inn
            ws2.cell(row=rr, column=4).value = float(bid.price)
            ws2.cell(row=rr, column=5).value = bid.created_at.strftime("%d.%m.%Y %H:%M")
            ws2.cell(row=rr, column=6).value = bid.created_by.get_full_name() if bid.created_by else '-'
            ws2.cell(row=rr, column=7).value = 'Да' if getattr(bid, 'is_active', False) else 'Нет'
            rr += 1

    # Настройка ширины столбцов для второго листа
    column_widths2 = [20, 30, 15, 15, 20, 25, 10]
    for col, width in enumerate(column_widths2, start=1):
        ws2.column_dimensions[get_column_letter(col)].width = width

    for r in range(2, rr):
        ws2.cell(row=r, column=4).number_format = '#,##0.00'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'tender_{tender.id}_{tender.name.replace(" ", "_")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def _build_tender_matrix_excel_response(tender):
    """
    Генерация Excel отчёта в матричном виде (Компании x Направления).
    Показывает лучшую ставку каждой компании по каждому направлению.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Матрица {tender.name[:20]}"

    # Стили
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")  # Черный фон
    header_font = Font(bold=True, color="FFFFFF", size=11)  # Белый текст
    border_style = openpyxl.styles.Side(border_style="thin", color="FFFFFF")
    white_border = openpyxl.styles.Border(top=border_style, left=border_style, right=border_style, bottom=border_style)
    
    # Данные
    directions = list(tender.directions.all().order_by('id'))
    # Получаем все компании, которые делали ставки в этом тендере
    companies_ids = Bid.objects.filter(direction__tender=tender).values_list('company', flat=True).distinct()
    companies = list(Company.objects.filter(id__in=companies_ids).order_by('name'))
    
    # 1. Заголовок
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(directions) + 1)
    ws['A1'] = f"{tender.name}"
    ws['A1'].font = Font(bold=True, color="FFFFFF", size=14)
    ws['A1'].fill = header_fill
    
    # 2. Шапка (Города)
    # A2 - пусто (или "Компания"), B2.. - Города
    ws.cell(row=2, column=1).value = ""  # Угловая ячейка
    ws.cell(row=2, column=1).fill = header_fill
    
    col_idx = 2
    for direction in directions:
        cell = ws.cell(row=2, column=col_idx)
        cell.value = direction.city_name
        cell.fill = header_fill
        cell.font = header_font
        cell.border = white_border
        col_idx += 1
        
    # 3. Строки (Компании)
    row_idx = 3
    for company in companies:
        # Имя компании
        c_cell = ws.cell(row=row_idx, column=1)
        c_cell.value = company.name
        c_cell.fill = header_fill # Компания тоже на черном фоне, как в примере
        c_cell.font = header_font
        c_cell.border = white_border
        
        col_idx = 2
        for direction in directions:
            # Ищем лучшую ставку этой компании по этому направлению
            best_bid = Bid.objects.filter(
                direction=direction, 
                company=company, 
                is_active=True
            ).order_by('price').first()
            
            val_cell = ws.cell(row=row_idx, column=col_idx)
            val_cell.fill = header_fill # Весь фон черный
            val_cell.font = header_font
            val_cell.border = white_border
            
            if best_bid:
                val_cell.value = float(best_bid.price)
                # Формат чисел (без .00 если целое, иначе с копейками) (хотя в примере целые)
                val_cell.number_format = '0' 
            else:
                val_cell.value = "-"
                val_cell.alignment = Alignment(horizontal='center')
            
            col_idx += 1
        row_idx += 1

    # Автоширина
    ws.column_dimensions['A'].width = 30
    for i in range(2, len(directions) + 2):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = 15

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'matrix_tender_{tender.id}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@staff_member_required
def download_tender_report(request, tender_id):
    """Генерирует Excel-файл с результатами тендера"""
    tender = get_object_or_404(Tender, id=tender_id)
    return _build_tender_excel_response(tender)


def _build_tender_protocol_pdf_response(tender):
    """
    Генерация PDF протокола тендера.
    Требует установленный пакет reportlab.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    except ImportError as e:
        raise ImportError("Не установлен reportlab. Установите: pip install reportlab") from e

    import io
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, title=f"Протокол тендера {tender.name}")
    styles = getSampleStyleSheet()

    story = []
    story.append(Paragraph(f"Протокол тендера: <b>{tender.name}</b>", styles["Title"]))
    story.append(Spacer(1, 10))
    story.append(Paragraph(f"Статус: {tender.get_status_display()}", styles["Normal"]))
    story.append(Paragraph(f"Дата создания: {tender.created_at.strftime('%d.%m.%Y %H:%M')}", styles["Normal"]))
    if tender.end_time:
        story.append(Paragraph(f"Дата закрытия: {tender.end_time.strftime('%d.%m.%Y %H:%M')}", styles["Normal"]))
    story.append(Spacer(1, 14))

    data = [["Направление", "Объём", "Старт", "Финал", "Победитель"]]
    for d in tender.directions.all():
        data.append([
            d.city_name,
            str(d.volume),
            f"{d.start_price} ₽",
            f"{d.final_price} ₽" if d.final_price else "-",
            d.winner.name if d.winner else "-",
        ])

    table = Table(data, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#366092")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(table)

    doc.build(story)
    pdf = buf.getvalue()
    buf.close()

    response = HttpResponse(content_type="application/pdf")
    filename = f"protocol_tender_{tender.id}_{tender.name.replace(' ', '_')}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response.write(pdf)
    return response


@staff_member_required
def download_tender_protocol_pdf(request, tender_id):
    """Скачать PDF протокол тендера (для админов)."""
    tender = get_object_or_404(Tender, id=tender_id)
    try:
        log_action(request.user, 'report_generated', 'tender', tender.id, {'type': 'pdf'}, request)
        return _build_tender_protocol_pdf_response(tender)
    except ImportError as e:
        messages.error(request, str(e))
        return redirect('admin:core_tender_changelist')

    